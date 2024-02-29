import os
import pandas as pd
import locale
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
import threading
from time import sleep
import re
from colorama import Fore
class MergeExcelReports:
     

    def formatar_cnpj(self, cnpj):
                # Verifica se o CNPJ é uma string válida
                if isinstance(cnpj, str) and len(cnpj) == 14 and cnpj.isdigit():
                    # Aplica a formatação
                    cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
                    return cnpj_formatado
                else:
                    return cnpj
                

    # função que remove arquivo do diretório de entrada com o nome "consolidado_final.xlsx"
    def remove_file(self, folder_path):
        try:
            # remove todo arquivo que inicia com "CONSOLIDADO"
            for file in os.listdir(folder_path):
                if file.startswith('CONSOLIDADO'):
                    os.remove(os.path.join(folder_path, file))
                    sleep(0.5)
                    
        except FileNotFoundError:
            pass                

    # Função para formatar valores com o símbolo da moeda
    def format_currency(self, value):
        return locale.currency(value, grouping=True, symbol='R$')
    
        
    
    def corrigir_valor_faturamento(self,valor):
        try:
            if pd.notna(valor):
                # Remover pontos e substituir vírgula por ponto
                valor = str(valor).replace('.', '').replace(',', '.')
                # Certificar-se de que há apenas um ponto decimal
                if valor.count('.') > 1:
                    valor = valor.replace('.', '', valor.count('.') - 1)
                return float(valor)
            else:
                return valor
        except Exception as e:
            print(f"Erro ao processar valor: {valor}, Erro: {e}")
            return None




    def merge_excel_reports(self, folder_path, output_folder):
        """Mescla arquivos Excel em um único arquivo Excel."""

        # remove arquivo do diretório de entrada com o nome "consolidado_final.xlsx"
        self.remove_file(folder_path)
        sleep(0.5)

        try:
            files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]
            
        except FileNotFoundError:
            print(Fore.RED + f'Nenhum arquivo encontrado em {folder_path}' + Fore.RESET)

        # Verifica se há apenas um arquivo no diretorio
        if len(files) == 1:
            print(f'Apenas um arquivo encontrado em {folder_path}')
            return
        
        # Dataframe vazio para armazenar os dados mesclados
        merged_data = pd.DataFrame()       

      
        # iteração sobre os arquivos Excel
        for file in files:
            file_path = os.path.join(folder_path, file)
            try:
                df = pd.read_excel(file_path, engine='openpyxl', sheet_name='CONSOLIDADO')
            except Exception as e:
                print(f'Erro ao ler o arquivo {file_path}: {e}')
                continue

            merged_data = merged_data._append(df, ignore_index=True )
            #print(f'Arquivo {file} adicionado ao consolidado com sucesso!')
            
            # formata células com cnpj para o formato xx.xxx.xxx/xxxx-xx
            cols_cnpj = ['CNPJ DO CLIENTE', 'CNPJ DE FATURAMENTO']

            # formatação da coluna CNPJ
            for col in cols_cnpj:
                # verifica se a coluna é do tipo string
                if pd.api.types.is_string_dtype(merged_data[col]):
                    merged_data[col] = merged_data[col].str.replace(r'(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5')
                else:
                    try:
                        # converte a coluna para string
                        merged_data[col] = merged_data[col].astype(str).str.replace(r'(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5')
                    except Exception as e:
                        print(f'Erro ao formatar a coluna {col}: {e}')
                        
       
        if merged_data.empty:
            print(f'Nenhum dado encontrado em {folder_path}')
            return

        # nome do arquivo de saída com o nome do cliente
        file_name = merged_data['NOME DO CLIENTE'].iloc[0]
        # remove caracteres especiais do nome do cliente
        file_name = file_name.replace(' ', '_').replace('/', '_').replace('-', '_').replace('.', '_')

        # Limitar o comprimento do nome do arquivo
        max_filename_length = 255
        file_name = file_name[:max_filename_length].strip()

        # Caminho completo para o arquivo de saída
        os.makedirs(os.path.dirname(output_folder), exist_ok=True)  # Verifica se o diretório pai existe, se não, cria-o
        # Caminho completo para o arquivo de saída
        output_file = os.path.join(output_folder, f'CONSOLIDADO_{file_name}.xlsx')

        # Verifica se existe algum arquivo com o mesmo nome, se existir, ignora próximo passo
        if os.path.exists(output_file):
            print(f'Arquivo {output_file} já existe!')
            return

        # Aplicar a lógica de conversão na coluna 'VLR TOTAL FATURAMENTO'
        #merged_data['VLR TOTAL FATURAMENTO'] = merged_data['VLR TOTAL FATURAMENTO'].apply(self.corrigir_valor_faturamento)
        #merged_data['VALOR TOTAL GERADO'] = merged_data['VALOR TOTAL GERADO'].apply(self.corrigir_valor_faturamento)
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                merged_data.to_excel(writer, sheet_name="RELATÓRIO", index=False, engine='openpyxl')

                # agrupar os valores das colunas "VALOR TOTAL GERADO" e "VALOR TOTAL FATURADO" por "PROJETO", "OBRA" e "CONTRATO LEGADO"
                sintese_df = merged_data.groupby(['PROJETO', 'OBRA', 'CONTRATO LEGADO'], as_index=False).agg(
                    {'VALOR TOTAL GERADO': 'sum', 'VLR TOTAL FATURAMENTO': 'sum'})
                # renomear as colunas
                sintese_df = sintese_df.rename(columns={'VLR TOTAL FATURAMENTO': 'VALOR TOTAL FATURADO'})
                

                # formatação da planilha "CONSOLIDADO"
                worksheet = writer.sheets['RELATÓRIO']
                for column in range(1, worksheet.max_column + 1):
                    worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = 20
                    worksheet.cell(row=1, column=column).font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                    worksheet.cell(row=1, column=column).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    # alinhar o texto no meio
                    worksheet.cell(row=1, column=column).alignment = Alignment(horizontal='center', vertical='center', )
                    worksheet.row_dimensions[1].height = 24

                # Configuração para o formato brasileiro
                locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

                sintese_df.to_excel(writer, sheet_name='SÍNTESE', index=False)

                # Adiciona "TOTAL" abaixo da célula "C"
                worksheet = writer.sheets['SÍNTESE']
                worksheet.cell(row=worksheet.max_row + 2, column=4, value='TOTAL')

                # negrito na célula "TOTAL"
                worksheet.cell(row=worksheet.max_row, column=4).font = Font(bold=True)
                worksheet.cell(row=worksheet.max_row, column=5).font = Font(bold=True)

                # Soma os valores da coluna "D" (VALOR TOTAL FATURADO)
                total_valor_a_cobrar = sintese_df['VALOR TOTAL GERADO'].sum()
                
                # Soma os valores da coluna "E" (VALOR TOTAL FATURADO)
                total_valor_total_previo = sintese_df['VALOR TOTAL FATURADO'].sum()


                # negrito na célula "VALOR TOTAL FATURADO"
                worksheet.cell(row=worksheet.max_row, column=4).font = Font(bold=True)

                # formatação da soma dos valores da coluna "D, E"
                total_valor_a_cobrar = locale.currency(total_valor_a_cobrar, grouping=True)
                total_valor_total_previo = locale.currency(total_valor_total_previo, grouping=True)

                worksheet.cell(row=worksheet.max_row, column=4, value=total_valor_a_cobrar)
                worksheet.cell(row=worksheet.max_row, column=5, value=total_valor_total_previo)

                # Aplicar cor vermelha ao cabeçalho das colunas A, B, C e D e negrito e tipografia "Alwyn New Light"
                for column in 'ABCDE':
                    header_cell = worksheet[f"{column}1"]
                    header_cell.font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                    header_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    header_cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                    # alinhar o texto na esquerda
                    header_cell.alignment = Alignment(horizontal='left', vertical='center', )

                # adicona bordas externas à planilha "SÍNTESE"
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

                # formatar largura das colunas
                writer.sheets['SÍNTESE'].column_dimensions['A'].width = 20
                writer.sheets['SÍNTESE'].column_dimensions['B'].width = 15
                writer.sheets['SÍNTESE'].column_dimensions['C'].width = 31
                writer.sheets['SÍNTESE'].column_dimensions['D'].width = 23
                writer.sheets['SÍNTESE'].column_dimensions['E'].width = 23
        
        except Exception as e:
            print(f'Erro ao mesclar arquivos Excel: {e}')
            return

          
            