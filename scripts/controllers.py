import pandas as pd 
import time
import os
import glob
import re
import unidecode
from colorama import Fore
import sys
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from database import ConnectPostgresQL, OrdersTable
from sqlalchemy.exc import IntegrityError
import logging
import shutil 
import re
import locale
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
#import pyspark.pandas as pd
from fuzzywuzzy import process
from streamlit import file_uploader
from time import sleep



class FinalReport:

    def __init__(self, host):
        # Crie uma instância de ConnectPostgresQL usando o host do seu banco de dados PostgreSQL
        self.db_connection = ConnectPostgresQL(host)
        self.session = self.db_connection.Session()

    """ Feche a sessão ao destruir a instância da classe """
    def __del__(self):
        
        self.session.close()
        self.db_connection.connect().close()

    # função para padronizar nomes das colunas        
    def padronizar_nomes_colunas(df):
        df.columns = (
            df.columns
            .str.lower()  # Converte para minúsculas
            .str.replace(r'[^a-zA-Z0-9_]', '_', regex=True)  # Substitui caracteres especiais por underscores
        )

    # função para formatar cnpj
    def formatar_cnpj(self, cnpj):
        # Verifica se o CNPJ é uma string válida
        if isinstance(cnpj, str) and len(cnpj) == 14 and cnpj.isdigit():
            # Aplica a formatação
            cnpj_formatado = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            return cnpj_formatado
        else:
            return cnpj
    
    # função para formatar as células do arquivo final
    def format_cells(self, path):
        """ função esta sendo usada na função rename_format_columns"""

        # define o diretório
        directory = path
        # percorre o diretório e localiza os arquivos excel
        for filename in os.listdir(directory):
            if filename.endswith('.xlsx'):
                # caminho completo do arquivo
                file_path = os.path.join(directory, filename)
                # Lê o arquivo
                df = pd.read_excel(file_path, sheet_name='CONSOLIDADO', engine='openpyxl')
                
                # seleciona as colunas que serão formatadas
                cnpj_cols = []
                try:
                    # itera sobre as colunas e aplica a função formatar_cnpj
                    for col in cnpj_cols:
                        df[col] = df[col].apply(self.formatar_cnpj)

                    # itera sobre as colunas e aplica a função strip e strftime    
                    for col in df.columns:
                        # Verifica se a coluna é do tipo string antes de aplicar .str
                        if pd.api.types.is_string_dtype(df[col]):
                            df[col] = df[col].str.strip()
                        elif pd.api.types.is_numeric_dtype(df[col]):
                            # Converte colunas numéricas para string antes de usar .str
                            df[col] = df[col].astype(str).str.strip()

                        # Verifica se a coluna é do tipo datetime antes de formatar
                        if pd.api.types.is_datetime64_any_dtype(df[col]):
                            df[col] = df[col].dt.strftime('%d/%m/%Y')                
                        # Retorna o DataFrame modificado
                        print(f'Renomeando coluna dataframes ...')
                        return df
                    
                # caso ocorra algum erro, exibe o erro    
                except Exception as e:
                    print(f"Erro ao formatar colunas: {e}")
                    return None
                    
                
    # verifica se o pedido existe no banco de dados
    def does_order_exist(self, order_number):
        query = f'SELECT 1 FROM {OrdersTable.__tablename__} WHERE {OrdersTable.__tablename__}.pedido_faturamento = {order_number} LIMIT 1'
        result = self.db_connection.engine.execute(query)
        return result.scalar() is not None


    """ Função para checar novos pedidos e atualizar o banco de dados"""
    def check_and_update_orders(self, extractor_file_path, col):
        start = time.time()
        """Método para verificar e atualizar pedidos ausentes no banco de dados"""

        #print(Fore.LIGHTYELLOW_EX + 'Verificando novos pedidos e atualizando o banco de dados...\n' + Fore.RESET)

        try:
            for filename in os.listdir(extractor_file_path):
                if filename.endswith('.xlsx') and not filename.startswith('~$'):
                    # caminho completo do arquivo
                    file_path = os.path.join(extractor_file_path, filename)
                    # Carrega o arquivo e verifica extrator TOTVS com os pedidos
                    extract_df = pd.read_excel(file_path, sheet_name='2-Resultado', engine='openpyxl', header=1)

                    # verifica se a coluna "Nome do Cliente" esta presente no indice 1(2ª linha)
                    if 'Pedido Faturamento' in extract_df.iloc[1].values:
                        extract_df.columns = extract_df.iloc[1]                   

                    # Padroniza o nome da coluna para minúsculas e substitui espaços por underscore
                    extract_df.columns = extract_df.columns.str.lower().str.replace(' ', '_').str.replace('.', '') \
                        .str.replace('-', '') \
                        .str.replace('ç', 'c') \
                        .str.replace('cálculo_reajuste', 'calculo_reajuste')  # Remove o acento "´"

                    # Verifica se a coluna existe no arquivo
                    col_lower = col.lower().replace(' ', '_')
                    if col_lower not in extract_df.columns:
                        print(f'Coluna {col} não encontrada no arquivo')
                        return
                        
                    # Verifica se há células vazias na coluna PEDIDO e dropa as linhas
                    extract_df = extract_df.dropna(subset=[col_lower])
                    
                  
                    # Converte a coluna PEDIDO para numérico
                    extract_df[col_lower] = pd.to_numeric(extract_df[col_lower], errors='coerce')
                    print(f'Total de registros no extrator: {len(extract_df)} | Arquivo {filename}')

                    # Filtra o DataFrame para incluir apenas os pedidos mais recentes
                    extract_df = extract_df[extract_df[col_lower] >= 0]

                    # trata o possivel erro invalid literal for int() with base 10
                    extract_df = extract_df[extract_df[col_lower].notna()]
                    extract_df[col_lower] = extract_df[col_lower].astype(int)                    

                    # Carrega os pedidos já existentes no banco de dados, convertendo a coluna para inteiro
                    existing_orders = set(int(order) for order in pd.read_sql_query(
                        f'SELECT DISTINCT {col} FROM {OrdersTable.__tablename__}', self.db_connection.engine)[col])

                    # Identifica os pedidos ausentes
                    new_orders = set(extract_df[col_lower]) - existing_orders
                    print(f'Total de novos pedidos no extrator: {len(new_orders)} | Arquivo {filename}')
                    
                    
                    if len(new_orders) > 0:
                        # Reinicializa a variável new_orders_df
                        new_orders_df = pd.DataFrame()

                        # Cria um DataFrame apenas com os pedidos ausentes
                        new_orders_df = extract_df[extract_df[col_lower].isin(new_orders)].copy()

                        # Verifica se há novos pedidos antes de continuar
                        if not new_orders_df.empty:
                            # caminho do diretório NOVOS_PEDIDOS
                            #path = r'/home/administrator/WindowsShare/01 - FATURAMENTO/03 - DATA_RAW'
                            path = r'C:\DataWare\data\consolidated_files\consolidated_validated\NOVOS_PEDIDOS'
                            # cria o diretório NOVOS_PEDIDOS se não existir
                            os.makedirs(path, exist_ok=True)
                            # percorre o DataFrame agrupando os pedidos por cliente
                            for order_number, order_group in new_orders_df.groupby(col_lower):
                                # remove caracteres inválidos do nome do cliente e cria o nome do arquivo
                                client_name_valid = order_group['nome_do_cliente'].iloc[0].translate(
                                    str.maketrans('', '', r'\/:*?"<>|'))
                                # define o nome e cria o arquivo
                                file_name = f'{order_number}_{client_name_valid}.xlsx'
                                # caminho completo do arquivo para salvar
                                file_path = os.path.join(path, file_name)
                                # salva o arquivo em excel
                                order_group.to_excel(file_path, sheet_name='CONSOLIDADO', index=False, engine='openpyxl')
                                print(f'Novo arquivo {file_name} criado.')

                            # Atualiza o banco de dados com os pedidos ausentes
                            print('Atualizando banco de dados....!')
                            try:
                                new_orders_df = new_orders_df.rename(columns={'cálculo_reajuste': 'calculo_reajuste'})

                                new_orders_df.to_sql(OrdersTable.__tablename__, self.db_connection.engine, if_exists='append', index=False, method='multi')
                                print(Fore.GREEN + 'Banco de dados atualizado com novos pedidos' + Fore.RESET)
                            except IntegrityError as e:
                                print('Erro ao atualizar o banco de dados:', e)

                            # pula o processamento dos clientes abaixo (grandes clientes)
                            special_clients = ['teste']

                            def save_order_excel(order, project):
                                order_df = extract_df[extract_df[col_lower] == order]
                                
                                if not order_df.empty:
                                    client_name = order_df['CLIENTE'].iloc[0]
                                    
                                    if client_name in special_clients:
                                        print(f'Relatório {client_name} será gerado manualmente')
                                        return

                                    client_name_safe = re.sub(r'[^a-zA-Z0-9_]', '_', unidecode.unidecode(client_name))
                                    sheet_names = ['LAVORO', 'CONSOLIDADO']

                                    # itera sobre as sheets do arquivo
                                    for sheet in sheet_names:
                                        # Define o nome do arquivo no padrão desejado, incluindo o número do projeto
                                        file_name = f'{project}_{order}_{client_name_safe}.xlsx'
                                        # Caminho completo do arquivo para salvar
                                        file_path = os.path.join(path, file_name)
                                        # Salva o arquivo em excel
                                        order_df.to_excel(file_path, sheet_name=sheet, index=False, engine='openpyxl')
                                        print(f'Arquivo {file_name} criado.')

                            with ThreadPoolExecutor() as executor:
                                executor.map(lambda order: save_order_excel(order, 'projeto'), new_orders)

                            print(f'Pedidos salvos no diretório NOVOS_PEDIDOS')
                            print('Verificação e atualização concluídas.\n')
                            end = time.time()
                            #print(f'Tempo de execução do código: {end - start}')

                        else:
                            print('Nenhum pedido novo encontrado.\n')

        
        except PermissionError as e:
            print(f"Erro de permissão ao acessar {extractor_file_path}: {e}")
            print('Corrija as permissões e tente novamente.')
            return False
        except Exception as e:
            print(f"Erro ao processar arquivo: {e}")
            pass

           
    """ Função que cria um arquivo único do cliente com todos os pedidos"""
    def merge_same_client(self, news_orders, output_path):
        # Inicia o contador de tempo de execução do método
        start_time = time.time()
        # lista os arquivos do diretório

        xlsx_files = glob.glob(os.path.join(news_orders, '*.xlsx'))
        # cria um dataframe vazio
        combined_df = pd.DataFrame()

        # itera sobre os arquivos do diretório
        for file in xlsx_files:
            # carrega o arquivo
            df = pd.read_excel(file, sheet_name='CONSOLIDADO')
            # concatena o dataframe do arquivo com o dataframe combinado
            combined_df = pd.concat([combined_df, df], ignore_index=True)

        # Salva o dataframe combinado em um arquivo excel
        combined_df.to_excel(output_path, sheet_name='CONSOLIDADO', engine='openpyxl', index=False)
        end_time = time.time()
        print(f'Tempo de execução: {end_time - start_time}')

    
    
    def corrigir_valor_faturamento(self,valor):
        try:
            if pd.notna(valor):
                # Remover pontos e substituir vírgula por ponto
                valor = str(valor).replace('.', '').replace(',', '.')
                # Certificar-se de que há apenas um ponto decimal
                if valor.count('.') > 1:
                    valor = valor.replace('.', '', valor.count('.') - 1)
                return float(valor)
            else:
                return valor
        except Exception as e:
            print(f"Erro ao processar valor: {valor}, Erro: {e}")
            return None
    
    
    """ Função que renomeia e formata o arquivo final com cores da Arklok"""
    def rename_format_columns(self, directory):
        
        # dicionário com os nomes das colunas (chave) e os nomes formatados (valor
        new_names = {
            'codigo_cliente': 'CÓDIGO CLIENTE',
            'loja_cliente': 'LOJA CLIENTE',
            'nome_do_cliente': 'NOME DO CLIENTE',
            'cnpj_do_cliente': 'CNPJ DO CLIENTE',
            'email': 'E-MAIL',
            'contrato_legado': 'CONTRATO LEGADO',
            'projeto': 'PROJETO',
            'obra': 'OBRA',
            'nome_da_obra': 'NOME DA OBRA',
            'numero_da_as': 'NUMERO DA AS',
            'pedido_de_remessa': 'PEDIDO DE REMESSA',
            'nota_de_remessa': 'NOTA DE REMESSA',
            'serie_da_nf_remessa': 'SERIE DA NF REMESSA',
            'data_de_remessa': 'DATA DE REMESSA',
            'cnpj_de_remessa': 'CNPJ DE REMESSA',
            'id_equipamento': 'ID EQUIPAMENTO',
            'id_equip_substituido': 'ID EQUIP SUBSTITUIDO',
            'data_da_substituicao': 'DATA DA SUBSTITUICAO',
            'equipamento': 'EQUIPAMENTO',
            'tipo_de_servico': 'TIPO DE SERVICO',
            'tipo_de_operacao': 'TIPO DE OPERACAO',
            'produto': 'PRODUTO',
            'descricao_do_produto': 'DESCRIÇÃO DO PRODUTO',
            'quantidade': 'QUANTIDADE',
            'valor_de_origem': 'VALOR DE ORIGEM',
            'valor_unitario': 'VALOR UNITÁRIO',
            'valor_bruto': 'VALOR BRUTO',
            'desconto': 'DESCONTO',
            'acrescimo': 'ACRÉSCIMO',
            'data_de_ativacao_legado': 'DATA DE ATIVAÇÃO LEGADO',
            'data_de_ativacao': 'DATA DE ATIVAÇÃO',
            'ultimo_faturamento': 'ÚLTIMO FATURAMENTO',
            'data_proximo_faturamento': 'DATA PRÓXIMO FATURAMENTO',
            'data_fim_locacao': 'DATA FIM LOCACAO',
            'dias_de_locacao': 'DIAS DE LOCAÇÃO',
            'prazo_do_contrato': 'PRAZO DO CONTRATO',
            'previsao_retirada': 'PREVISÃO RETIRADA',
            'solicitacao_retirada': 'SOLICITAÇÃO RETIRADA',
            'tipo_do_mes': 'TIPO DO MÊS',
            'mes_fixo': 'MÊS FIXO',
            'data_base_reajuste': 'DATA BASE REAJUSTE',
            'indexador': 'INDEXADOR',
            'data_do_reajuste': 'DATA DO REAJUSTE',
            'indice_aplicado': 'ÍNDICE APLICADO',
            'calculo_reajuste': 'CÁLCULO REAJUSTE',
            'franquia': 'FRANQUIA',
            'class_faturaento': 'CLASS FATURAMENTO',
            'cobra': 'COBRA ?',
            'data_entrada': 'DATA ENTRADA',
            'centro_de_custos': 'CENTRO DE CUSTOS',
            'pedido_faturamento': 'PEDIDO FATURAMENTO',
            'emissao_pedido': 'EMISSÃO PEDIDO',
            'qtde_pedido': 'QTDE PEDIDO',   
            'vlr_unitario_pedido': 'VLR UNITÁRIO PEDIDO',
            'vlr_total_pedido': 'VALOR TOTAL GERADO',
            'percent_desconto': 'PERCENT DESCONTO',
            'vlr_desconto': 'VLR DESCONTO',
            'tes': 'TES',
            'natureza': 'NATUREZA',
            'nf_de_faturamento': 'NF DE FATURAMENTO',
            'serie_de_faturamento': 'SERIE DE FATURAMENTO',
            'data_de_faturamento': 'DATA DE FATURAMENTO',
            'cliente_faturamento': 'CLIENTE FATURAMENTO',
            'loja_faturameto' : 'LOJA FATURAMENTO',
            'nome_cli_faturamento': 'NOME CLI FATURAMENTO',
            'cnpj_de_faturamento': 'CNPJ DE FATURAMENTO',
            'qtde_faturamento': 'QTDE FATURAMENTO',
            'vlr_unitario_faturamento': 'VLR UNITÁRIO FATURAMENTO',
            'vlr_total_faturamento': 'VLR TOTAL FATURAMENTO',
            'periodo_de_faturamento': 'PERÍODO DE FATURAMENTO',
            'origem_do_dado': 'ORIGEM DO DADO',
            'serie_do_equipamento': 'SERIE DO EQUIPAMENTO',

        }
        
        try:
            # percorre o diretório e localiza os arquivos excel 
            for filename in os.listdir(directory):
                if filename.endswith('.xlsx') and not filename.startswith('~$'):
                    # caminho completo do arquivo
                    file_path = os.path.join(directory, filename)
                    # Lê o arquivo
                    df = pd.read_excel(file_path, sheet_name='CONSOLIDADO', engine='openpyxl')                   
                    
                    # renomeia as colunas
                    df = df.rename(columns=new_names)

                    # dropas as colunas que não serão usadas
                    columns_to_drop = ['E-MAIL', 'NOME DA OBRA', 'NUMERO DA AS', 'PEDIDO DE REMESSA', 'NOTA DE REMESSA', 'SERIE DA NF REMESSA',
                                       'DATA DE REMESSA', 'CNPJ DE REMESSA', 'ID EQUIP SUBSTITUIDO', 'DATA DA SUBSTITUICAO', 'TIPO DE SERVICO',
                                       'TIPO DE OPERACAO', 'PRODUTO', 'DESCONTO', 'ÚLTIMO FATURAMENTO', 'DATA PRÓXIMO FATURAMENTO', 'DATA FIM LOCACAO',
                                       'PRAZO DO CONTRATO', 'PREVISÃO RETIRADA', 'SOLICITAÇÃO RETIRADA', 'TIPO DO MÊS', 'MÊS FIXO',
                                       'DATA DO REAJUSTE', 'FRANQUIA', 'CLASS FATURAMENTO', 'COBRA ?', 'DATA ENTRADA', 'CENTRO DE CUSTOS', 
                                       'PEDIDO FATURAMENTO', 'EMISSÃO PEDIDO', 'QTDE PEDIDO', 'VLR UNITÁRIO PEDIDO', 'PERCENT DESCONTO', 
                                       'VLR DESCONTO', 'TES', 'NATUREZA', 'SERIE DE FATURAMENTO', 'VLR UNITÁRIO FATURAMENTO', 
                                       'CLIENTE FATURAMENTO', 'LOJA FATURAMENTO', 'NOME CLI FATURAMENTO', 'QTDE FATURAMENTO', 
                                       'ORIGEM DO DADO', 'SERIE DO EQUIPAMENTO']
                    

                    df = df.drop(columns_to_drop, axis=1, errors='ignore')

                    # reordena as colunas
                    df = df[['CÓDIGO CLIENTE', 'NOME DO CLIENTE', 'LOJA CLIENTE', 'CNPJ DO CLIENTE', 'CNPJ DE FATURAMENTO',
                            'PROJETO', 'OBRA', 'ID EQUIPAMENTO', 'EQUIPAMENTO', 'DESCRIÇÃO DO PRODUTO', 'DATA DE ATIVAÇÃO LEGADO', 
                            'PERÍODO DE FATURAMENTO', 'DIAS DE LOCAÇÃO', 'VALOR UNITÁRIO', 'VALOR BRUTO', 'DATA DE ATIVAÇÃO',
                            'QUANTIDADE', 'VALOR TOTAL GERADO', 'VLR TOTAL FATURAMENTO', 
                            'NF DE FATURAMENTO',  'DATA DE FATURAMENTO', 'DATA BASE REAJUSTE', 'VALOR DE ORIGEM', 
                            'INDEXADOR', 'CÁLCULO REAJUSTE', 'ÍNDICE APLICADO', 'ACRÉSCIMO', 'CONTRATO LEGADO']]
                     
                    
                    # formata células com datas para o formato dd/mm/aaaa
                    cols_date = ['DATA DE ATIVAÇÃO', 'DATA DE ATIVAÇÃO LEGADO', 'DATA BASE REAJUSTE', 'DATA DE FATURAMENTO']
                    
                    # itera sobre as colunas 
                    for col in cols_date:
                        # Verifica se a coluna é do tipo datetime antes de formatar
                        if pd.api.types.is_datetime64_any_dtype(df[col]):
                            df[col] = df[col].dt.strftime('%d/%m/%Y')
                        else:
                            try:
                                # converte a coluna para datetime
                                df[col] = pd.to_datetime(df[col].loc[df[col].notna()], format='%d/%m/%Y', errors='coerce')
                            except Exception as e:
                                print(f'Erro ao converter data: {e}')
                    
                    # formata células com cnpj para o formato xx.xxx.xxx/xxxx-xx
                    cols_cnpj = ['CNPJ DO CLIENTE', 'CNPJ DE FATURAMENTO']
                    # itera sobre as colunas
                    for col in cols_cnpj:
                        # Verifica se a coluna é do tipo string antes de aplicar .str
                        if pd.api.types.is_string_dtype(df[col]):
                            # defini o formato do cnpj xx.xxx.xxx/xxxx-xx
                            df[col] = df[col].str.replace(r'(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5')
                        else:
                            try:
                                # converte a coluna para string
                                df[col] = df[col].astype(str).str.replace(r'(\d{2})(\d{3})(\d{3})(\d{4})(\d{2})', r'\1.\2.\3/\4-\5')
                            except Exception as e:
                                print(f'Erro ao converter cnpj: {e}')          
                     
                                        
                    # Aplicar a lógica de conversão na coluna 'VLR TOTAL FATURAMENTO'
                    df['VLR TOTAL FATURAMENTO'] = df['VLR TOTAL FATURAMENTO'].apply(self.corrigir_valor_faturamento)
                    df['VALOR TOTAL GERADO'] = df['VALOR TOTAL GERADO'].apply(self.corrigir_valor_faturamento)
                    

                    # salva o arquivo em excel
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        df.fillna({'CONTRATO LEGADO': '0'}, inplace=True)
                        
                        df.to_excel(writer, sheet_name='CONSOLIDADO', index=False, engine='openpyxl')

                        # sintese_df = df.groupby(['PROJETO', 'OBRA', 'CONTRATO LEGADO'], as_index=False)['VLR TOTAL FATURAMENTO'].sum()
                        sintese_df = df.groupby(['PROJETO', 'OBRA', 'CONTRATO LEGADO'], as_index=False).agg({'VALOR TOTAL GERADO': 'sum', 'VLR TOTAL FATURAMENTO': 'sum'})

                        sintese_df = sintese_df.rename(columns={'VLR TOTAL FATURAMENTO': 'VALOR TOTAL FATURADO'})
                        # sintese_df = sintese_df.rename(columns={'VALOR TOTAL PEDIDO': 'VALOR TOTAL GERADO'})

                        # formatação da planilha "CONSOLIDADO"
                        worksheet = writer.sheets['CONSOLIDADO']
                        for column in range(1, worksheet.max_column + 1):
                            worksheet.column_dimensions[worksheet.cell(row=1, column=column).column_letter].width = 25
                            worksheet.cell(row=1, column=column).font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                            worksheet.cell(row=1, column=column).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            # alinhar o texto no meio
                            worksheet.cell(row=1, column=column).alignment = Alignment(horizontal='center', vertical='center', )

                            worksheet.row_dimensions[1].height = 24
                        
                                                
                        # Configuração para o formato brasileiro
                        locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

                        # formatação da coluna VALOR TOTAL FATURADO

                        sintese_df['VALOR TOTAL FATURADO'] = sintese_df['VALOR TOTAL FATURADO'].apply(lambda x: locale.currency(float(x) if pd.notna(x) else 0, grouping=True, symbol='R$'))
                        sintese_df['VALOR TOTAL GERADO'] = sintese_df['VALOR TOTAL GERADO'].apply(lambda x: locale.currency(float(x) if pd.notna(x) else 0, grouping=True, symbol='R$'))
                        sintese_df.to_excel(writer, sheet_name='SÍNTESE', index=False)


                        # Adiciona "TOTAL" abaixo da célula "C"
                        worksheet = writer.sheets['SÍNTESE']
                        worksheet.cell(row=worksheet.max_row + 2, column=4, value='TOTAL')

                        # negrito na célula "TOTAL"
                        worksheet.cell(row=worksheet.max_row, column=4).font = Font(bold=True)
                        worksheet.cell(row=worksheet.max_row, column=5).font = Font(bold=True)
                        

                        # Soma os valores da coluna "D" (VALOR TOTAL FATURADO)
                        total_valor_a_cobrar = sintese_df['VALOR TOTAL GERADO'].apply(lambda x: locale.atof(x.split()[1])).sum()
                        # Soma os valores da coluna "E" (VALOR TOTAL FATURADO)
                        total_valor_total_previo = sintese_df['VALOR TOTAL FATURADO'].apply(lambda x: locale.atof(x.split()[1])).sum()



                        # negrito na célula "VALOR TOTAL FATURADO"
                        worksheet.cell(row=worksheet.max_row, column=4).font = Font(bold=True)
                        
                        # formatação da soma dos valores da coluna "D, E"
                        total_valor_a_cobrar = locale.currency(total_valor_a_cobrar, grouping=True)
                        total_valor_total_previo = locale.currency(total_valor_total_previo, grouping=True)
                        
                        worksheet.cell(row=worksheet.max_row, column=4, value=total_valor_a_cobrar)
                        worksheet.cell(row=worksheet.max_row, column=5, value=total_valor_total_previo)

                        # Adiciona logotipo ao cabeçalho da sheet 'SÍNTESE'
                        #img = Image(logo_path)
                        #worksheet.add_image(img, 'A1')

                        
                        # Aplicar cor vermelha ao cabeçalho das colunas A, B, C e D e negrito e tipografia "Alwyn New Light"
                        for column in 'ABCDE':
                            header_cell = worksheet[f"{column}1"]
                            header_cell.font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
                            header_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            header_cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                            header_cell.alignment = Alignment(horizontal='center', vertical='center', )

                            # Ajusta a altura da linha do cabeçalho
                            worksheet.row_dimensions[1].height = 24
                        
                        # adicona bordas externas à planilha "SÍNTESE"
                        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                            for cell in row:
                                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                                

                        # formatar largura das colunas
                        writer.sheets['SÍNTESE'].column_dimensions['A'].width = 20
                        writer.sheets['SÍNTESE'].column_dimensions['B'].width = 15
                        writer.sheets['SÍNTESE'].column_dimensions['C'].width = 31
                        writer.sheets['SÍNTESE'].column_dimensions['D'].width = 23
                        writer.sheets['SÍNTESE'].column_dimensions['E'].width = 23           

                    #print(f'Arquivos formatados com sucesso em {file_path}')

        except PermissionError as e:
            print(f"O arquivo {filename} está aberto: {e}")
            print('Feche o arquivo manualmente e tente novamente.')
            return False
        #print(f'Colunas não usadas no relatório final removidas com sucesso')        

    
    """ Função para formatar células do arquivo final com datas no formato dd/mm/aaaa"""
    def format_date_cells(self, directory):
        # define o diretório
        directory = directory
        # percorre o diretório e localiza os arquivos excel
        for filename in os.listdir(directory):
            if filename.endswith('.xlsx'):
                # caminho completo do arquivo
                file_path = os.path.join(directory, filename)
                # Lê o arquivo
                df = pd.read_excel(file_path, sheet_name='CONSOLIDADO', engine='openpyxl')
                # lista com as colunas que serão formatadas
                date_cols = ['DATA DE ATIVAÇÃO', 'PERÍODO INICIAL',
                            'DATA PRÓXIMO FATURAMENTO']
                try:
                    # itera sobre as colunas e aplica a função formatar_cnpj
                    for col in date_cols:
                        df[col] = df[col].dt.strftime('%d/%m/%Y')
                except Exception as e:
                    print(f"Erro ao formatar colunas: {e}")
                    return None
                
           
    # função get_excel_files de um diretório
    def get_excel_files(self, folder_path):
        try:
            # lista todos os arquivos no diretório
            files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
            merged_data = pd.DataFrame()
        except FileNotFoundError as e:
            print(f'Arquivo não encontrado: {e}')
            return None

        # itera sobre cada arquivo no diretório
        for file in files:
            file_path = os.path.join(folder_path, file)
            df = pd.read_excel(file_path, engine='openpyxl')
            merged_data = merged_data._append(df, ignore_index=True)


    # função para formatar CNPJ de um dataframe
    def format_cnpj(self, cnpj):
        if isinstance(cnpj, str) and len(cnpj) == 14 and cnpj.isdigit():
            return f'{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}.{cnpj[12:]}'
        else:
            return cnpj

    
    # função para formatar CNPJ de uma coluna de um dataframe
    def format_cnpj_column(self, df, column):
        df[column] = df[column].apply(self.format_cnpj)


    # função para escrever um dataframe em um arquivo excel
    def write_to_excel(self, df, output_file, sheet_name):
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

            # agrupar os valores das colunas "VALOR TOTAL GERADO" e "VALOR TOTAL FATURADO" por "PROJETO", "OBRA" e "CONTRATO LEGADO"
            sintese_df = df.groupby(['PROJETO', 'OBRA', 'CONTRATO LEGADO'], as_index=False).agg({'VALOR TOTAL GERADO': 'sum', 'VLR TOTAL FATURAMENTO': 'sum'})
            sintese_df = sintese_df.rename(columns={'VLR TOTAL FATURAMENTO': 'VALOR TOTAL FATURADO'})
    
            # formatar largura das colunas
            writer.sheets['SÍNTESE'].column_dimensions['A'].width = 20
            writer.sheets['SÍNTESE'].column_dimensions['B'].width = 15
            writer.sheets['SÍNTESE'].column_dimensions['C'].width = 31
            writer.sheets['SÍNTESE'].column_dimensions['D'].width = 23
            writer.sheets['SÍNTESE'].column_dimensions['E'].width = 23
    
    
    # função para mesclar arquivos excel de um diretório
    def merge_excel_files(self, dataframe):
        return pd.concat(dataframe, ignore_index=True)
    
    
    # função para síntese
    def sintese_to_excel(self, dataframe, worksheet):
        dataframe.to_excel(worksheet, sheet_name='SÍNTESE', index=False)

           
    # função para adicionar totais na planilha SÍNTESE e personalizar
    def add_total_sintese(self, worksheet, sintese_df):
        for column in range(1, worksheet.max_column + 1):
            worksheet.cell(row=worksheet.max_row + 2, column=4, value='TOTAL')
            worksheet.cell(row=worksheet.max_row, column=4).font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
            worksheet.cell(row=worksheet.max_row, column=column).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

            locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

            total_valor_a_cobrar = sintese_df['VALOR TOTAL GERADO'].sum()
            total_valor_previo = sintese_df['VALOR TOTAL FATURADO'].sum()

            total_valor_a_cobrar = locale.currency(total_valor_a_cobrar, grouping=True)
            total_valor_previo = locale.currency(total_valor_previo, grouping=True)

            worksheet.cell(row=worksheet.max_row, column=4, value=total_valor_a_cobrar)
            worksheet.cell(row=worksheet.max_row, column=5, value=total_valor_previo)


    # função para formatar células de um arquivo excel
    def format_worksheet(worksheet):
        for column in 'ABCDE':
            header_cell = worksheet[f'{column}1']
            header_cell.font = Font(color='FFFFFF', bold=True, name='Lato Regular', size=10)
            header_cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            header_cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
                                        top=Side(border_style='thin'), bottom=Side(border_sytle='thin'))
    
            header_cell.alignment = Alignment(horizontal='left', vertical='center')    
    #

        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(border_style='thin'),
                                     right=Side(border_style='thin'),
                                     top=Side(border_style='thin'),
                                     bottom=Side(border_style='thin'))
    


# Classe para processar e listar arquivos do diretório
class FileProcessor:
    # Definindo atributos da classe
    def __init__(self, extractor_file_path, invoiced_orders, news_orders, output_merge_path):
        self.extractor_file_path = extractor_file_path
        self.invoiced_orders = invoiced_orders
        self.news_orders = news_orders
        self.output_merge_path = output_merge_path

    # Método para obter os arquivos
    def get_files(self, file_type='.xlsx'):
        # percorre o diretório e retorna uma lista de arquivos com o caminho completo usando list comprehension
        return [(root, file) for root, dirs, files in os.walk(self.news_orders) \
                for file in files if file.endswith(file_type    )]

    
    # Método para processar os arquivos
    def process_file_list(self, filo_info):
        
        # obtém o caminho completo do arquivo
        root, file = filo_info
        full_path = os.path.join(root, file)
        
        # cria uma lista vazia
        xlsx_files = []
        
        # verifica se o arquivo é um arquivo excel e não é um arquivo temporário
        if file.lower().endswith('.xlsx') \
            and not file.startswith('~$'):
            print(f'{Fore.LIGHTCYAN_EX}Arquivo encontrado em: {full_path}{Fore.RESET}')

            # limpa o buffer de saída    
            sys.stdout.flush() 
            # adiciona o arquivo na lista de arquivos
            xlsx_files.append(full_path)

            # Obtem informações do arquivo
            file_status = os.stat(full_path)
            file_size = file_status.st_size
            filename = os.path.basename(full_path)
            file_path = os.path.dirname(full_path)
            file_date_create = datetime.fromtimestamp(file_status.st_ctime).strftime('%d/%m/%Y %H:%M:%S')
            file_date_modified = datetime.fromtimestamp(file_status.st_mtime).strftime('%d/%m/%Y %H:%M:%S')
            full_path_file = os.path.join(file_path, filename)
            
            # retorna um dicionário com as informações do arquivo
            return {
                
                'FILENAME': filename,
                'FULL_PATH_FILE': full_path_file,
                'FILE_SIZE': file_size,
                'CREATE_DATE': file_date_create,
                'MODIFIED_DATE': file_date_modified,

            }
        
        else:
            return None
        
    
    # Método para processar os arquivos em paralelo (multithreading) 
    def process_files_in_parallel(self, file_infos):
        # Cria um pool de threads
        with ThreadPoolExecutor() as executor:            
            # Processa os arquivos em paralelo com o método map (função, iterável)
            results = executor.map(self.process_file_list, file_infos)
        # retorna os resultados usando list comprehension
        return [result for result in results if result is not None]
    
    
    # Método para listar os arquivos
    def list_all_files(self, output_folder):
        start_time = time.time()
        # verficar se o diretório existe
        if not os.path.exists(output_folder):
            print(f'A pasta {output_folder} não existe')
            # se não existir, cria o diretório
            os.makedirs(output_folder)
            print(f'Criando a pasta {output_folder}')
            return 
        
        # 
        try:
            # lista todos os arquivos no diretório
            file_infos = self.get_files()
            
            # processa os arquivos em paralelo
            file_list = self.process_files_in_parallel(file_infos)
            # itera sobre cada arquivo
            for file_info in file_infos:
                try:
                    processed_file = self.process_file_list(file_info)
                    # verifica se o arquivo foi processado
                    if processed_file is not None:
                        # adiciona o arquivo na lista de arquivos
                        file_list.append(processed_file)
                # caso ocorra algum erro, exibe o erro        
                except FileNotFoundError as e:
                    print(f'Arquivo não encontrado: {e}')
                    continue
            
            # Cria um DataFrame com os dados dos arquivos
            df = pd.DataFrame(file_list)
            # Salva o DataFrame em um arquivo excel
            df.to_excel(os.path.join(output_folder, f'NOVOS_PEDIDOS_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'), 
                        sheet_name='NOVOS_PEDIDOS', index=False, engine='openpyxl')

            # finaliza a contagem do tempo de execução
            end_time = time.time()
            elapsed_time = end_time - start_time
            print(f'Rotina Listagem novos pedidos finalizda! Tempo de execução: {elapsed_time}')

        # caso ocorra algum erro, exibe o erro
        except PermissionError as e:
            print(f'O arquivo {self.news_orders} está aberto. Feche o arquivo e tente novamente')
            return
        
        # caso ocorra algum erro, exibe o erro
        except Exception as e:
            print(f'Ocorreu um erro no arquivo {self.news_orders}: {e}')
            return False

    
    # Criar pastas no diretório H:\\
    def make_folders_clients(self, batch_totvs_path, extractor_path, sheet_name, col):
        # carrega o arquivo com os clientes
        df = pd.read_excel(extractor_path, sheet_name, engine='openpyxl')
        # cria a coluna com o caminho completo do diretório
        basedir = batch_totvs_path
        
        # itera sobre cada cliente
        for client in df[col]:
            # cria o caminho completo do diretório
            client_path = os.path.join(basedir, client)

            # define o nome da pasta
            
            
            # verifica se o diretório já existe
            if not os.path.exists(client_path):
                # se não existir, cria o diretório
                os.makedirs(client_path)
                print(f'Pasta {client} criada com sucesso!')
            else:
                print(f'Pasta {client} já existe!')

    
    # função para excluir todos os arquivos da pasta copied_files
    def delete_xlsx(self, files_path):
        """
        Deletes all .xlsx files in the specified directory.

        Args:
            files_path (str): The path to the directory containing the .xlsx files.

        Returns:
            bool: True if all files were successfully deleted, False otherwise.
        """
        
        # verifica se a pasta existe
        if not os.path.exists(files_path):
            print(f"A pasta {files_path} não existe.")
            return

        try:
            # lista todos os arquivos no diretório
            file_list = [f for f in os.listdir(files_path) if f.endswith('.xlsx') and not f.startswith('~$')]
            # itera sobre cada arquivo no diretório
            for file_name in file_list:
                # caminho completo do arquivo
                input_file_path = os.path.join(files_path, file_name)
                # verifica se o arquivo existe e tem permissões de leitura
                if os.path.isfile(input_file_path) and os.access(input_file_path, os.R_OK):
                    # exclui o arquivo
                    os.remove(input_file_path)
                else:
                    print(f"Arquivo {file_name} não encontrado ou permissão negada.")
            
            #print(f"Todos os arquivos excluídos com sucesso!")
            return True
        
        except PermissionError as e:
            print(f"O arquivo {files_path} está aberto: {e}")
            print('Delete o arquivo manualmente.')
            return False
        
        except Exception as e:
            print(f"Ocorreu um erro ao excluir os arquivos: {e}")
            return False
        
    
    # função para excluir todos os arquivos da pasta copied_files
    def delete_xml(self, files_path):
        #logging.info(f"EXCLUINDO ARQUIVOS XML...")
        # verifica se a pasta existe
        if not os.path.exists(files_path):
            print(f"A pasta {files_path} não existe.")
            return

        try:
            # lista todos os arquivos no diretório
            file_list = [f for f in os.listdir(files_path) if f.endswith('.xml') and not f.startswith('~$')]
            # itera sobre cada arquivo no diretório
            for file_name in file_list:
                # caminho completo do arquivo
                input_file_path = os.path.join(files_path, file_name)
                # verifica se o arquivo existe e tem permissões de leitura
                if os.path.isfile(input_file_path) and os.access(input_file_path, os.R_OK):
                    # exclui o arquivo
                    os.remove(input_file_path)
                    print(f"Arquivo {file_name} excluído com sucesso!")
                else:
                    print(f"Arquivo {file_name} não encontrado ou permissão negada.")
            
            #print(f"Todos os arquivos excluídos com sucesso!")
            return True
        
        except PermissionError as e:
            print(f"O arquivo {files_path} está aberto: {e}")
            print('Delete o arquivo manualmente.')
            return False
        
        except Exception as e:
            print(f"Ocorreu um erro ao excluir os arquivos: {e}")
            return False 
            
    
    # função para previsão de pastas no diretório H:\\ usando inteligência artificial
    def find_closest_match(self, client_name, target_base_directory):
        # Lista de diretórios de destino disponíveis
        target_directories = [d for d in os.listdir(target_base_directory)]
        
        # Encontra a correspondência mais próxima no diretório de destino usando fuzzywuzzy
        best_match, score = process.extractOne(client_name, target_directories)

        # Define um limite de confiança de 80%
        threshold = 50
        if best_match and score >= threshold:
            # Substituir caracteres inválidos ou limitar o tamanho do nome do diretório
            valid_directory_name = ''.join(char for char in best_match if char.isalnum() or char in [' ', '_'])
            return os.path.join(target_base_directory, valid_directory_name)
        else:
            # Se não houver correspondência próxima, crie um novo diretório
            valid_client_name = ''.join(char for char in client_name if char.isalnum() or char in [' ', '_'])
            new_directory = os.path.join(target_base_directory, valid_client_name)
            os.makedirs(new_directory, exist_ok=True)
            print(f'Diretório criado: {new_directory}')
            return new_directory


    # função para distribuir os arquivos por cliente
    def move_file_to_client_folder(self, source_directory, target_directory):
        path = source_directory

        try:
            for filename in os.listdir(path):
                if filename.endswith('.xlsx') and not filename.startswith('~$'):
                    full_source = os.path.join(path, filename)
                    print(f'Arquivo {filename} encontrado em {full_source}')

                    client_name = filename.split('_')[1].split('.')[0]
                    print(f'Cliente: {client_name}')

                    target_directory = self.find_closest_match(client_name, target_directory)
                    

                    if target_directory:
                        target_path = os.path.join(target_directory, client_name)
                        os.makedirs(target_path, exist_ok=True)

                        target_path_file = os.path.join(target_path, filename)

                        if os.path.exists(target_path_file):
                            print(f'Substituindo arquivo existente: {target_path_file}')
                            os.remove(target_path_file)

                        shutil.move(full_source, target_path)
                        print(f'Movendo arquivo {filename} para {target_path}...')
                    else:
                        print(f'Criando diretório para o cliente {client_name}')
                        target_path = os.path.join(target_directory, client_name)
                        os.makedirs(target_path, exist_ok=True)

                        target_path_file = os.path.join(target_path, filename)

                        shutil.move(full_source, target_path)
                        print(f'Movendo arquivo {filename} para {target_path}...')
                
        except PermissionError as e:
            print(f"O arquivo {source_directory} está aberto: {e}")
            print(f'Mova o arquivo manualmente para o diretório {target_path}')
            return False


    # função soma dos valores da coluna "VALOR BRUTO" de todos os arquivos do diretório
    def accurent_billing_value(self, directory):
        # lista os arquivos do diretório
        xlsx_files = glob.glob(os.path.join(directory, '*.xlsx'))
        # cria um dataframe vazio
        combined_df = pd.DataFrame()

        # itera sobre os arquivos do diretório
        for file in xlsx_files:
            # carrega o arquivo
            df = pd.read_excel(file, sheet_name='CONSOLIDADO', engine='openpyxl')
            # concatena o dataframe do arquivo com o dataframe combinado
            combined_df = pd.concat([combined_df, df], ignore_index=True)

        # Soma os valores da coluna "VALOR BRUTO"
        total_billing_value = combined_df['VALOR BRUTO'].sum()
        print(f'Total billing value: {total_billing_value}')

        # formatação da soma dos valores da coluna "VALOR BRUTO"
        #total_billing_value = locale.currency(total_billing_value, grouping=True)

        return total_billing_value

        
    # função para criar pastas no diretório no novo formato
    def make_new_folders(self, dataframe, sheet_name, engine, directory):
        # lê o dataframe
        df = pd.read_excel(dataframe, sheet_name=sheet_name, engine=engine)

        for index, row in df.iterrows():
            # seleciona a coluna 'Projeto' e 'Nome do Cliente'
            project = row['Projeto']
            client_name = row['Nome do Cliente']
            # define o nome da pasta
            folder_name = f'{client_name}'
            # define o caminho completo da pasta
            folder_path = os.path.join(directory, folder_name)
            # define a data atual para criar pasta do mês
            current_date = datetime.now()
            # formata a data atual para o formato mm-aaaa
            month_year = current_date.strftime('%m-%Y')
            # adiciona a subpasta do mês e ano
            folder_path_with_month_year = os.path.join(folder_path, month_year)

            # verifica se o diretório já existe
            if not os.path.exists(folder_path_with_month_year):
                # se não existir, cria o diretório
                os.makedirs(folder_path_with_month_year)
                print(f'Pasta {folder_path_with_month_year} criada com sucesso!')
            else:
                print(f'Pasta {folder_path_with_month_year} já existe!')

            # movimenta os arquivos para a subpasta do mês e ano
            self.move_files_to_month_subfolder(project, directory, month_year)
            

    # movimenta os arquivos para a subpasta do mês e ano
    def move_files_to_month_subfolder(self, directory_origin, target_directory):
        # obtém os arquivos xlsx no subdiretório principal
        files_to_move = [file for file in os.listdir(directory_origin) if file.endswith('.xlsx') and not file.startswith('~$')]
        
        # cria a subpasta do mês e ano
        current_date = datetime.now()
        # formata a data atual para o formato mm-aaaa
        month_year = current_date.strftime('%m-%Y')
        
        for file_to_move in files_to_move:
            # estabelece caminho completo do arquivo na origem
            current_file_path = os.path.join(directory_origin, file_to_move)
                        
            # extrai o nome do cliente do nome do arquivo
            client_name_start = file_to_move.find('_') + 1
            client_name_end = file_to_move.find('.', client_name_start)
            client_name = file_to_move[client_name_start:client_name_end]
            print(f'Nome do cliente: {client_name}')               

            # estabelece caminho completo do arquivo na destino
            
            current_file_path_with_month = os.path.join(target_directory, client_name, month_year)
            
            # cria o diretório para o arquivo ser movido
            if not os.path.exists(current_file_path_with_month):
                # se não existir, cria o diretório
                os.makedirs(current_file_path_with_month)
                print(f'Pasta {current_file_path_with_month} criada com sucesso!')    
                # move o arquivo para o diretório correspondente ao nome do cliente

            # caminho completo do arquivo de destino
            destination_file_path = os.path.join(current_file_path_with_month, file_to_move)
            
            # verifica se o arquivo já existe e remove o arquivo no caso positivo
            if os.path.exists(destination_file_path):
                print(f'Arquivo {file_to_move} já existe no diretório {current_file_path_with_month}')
                os.remove(destination_file_path)
            
            # move o arquivo para o diretório correspondente ao nome do cliente no caso negativo
            shutil.move(current_file_path, current_file_path_with_month)
            print(f'Arquivo {file_to_move} movido para {current_file_path_with_month}')
                      

    # função para mover arquivos para diretório de processados simples
    def move_files_to_processed_folder(self, directory_origin, target_directory):
        # obtém os arquivos xlsx no subdiretório principal
        files_to_move = [file for file in os.listdir(directory_origin) if file.endswith('.xlsx') and not file.startswith('~$')]
        
        for file_to_move in files_to_move:
            # estabelece caminho completo do arquivo na origem
            current_file_path_origin = os.path.join(directory_origin, file_to_move)
            # estabelece caminho completo do arquivo na destino
            current_file_path_target = os.path.join(target_directory, file_to_move)
            # move o arquivo para o diretório correspondente ao nome do cliente se ele não estiver aberto
            if current_file_path_origin.startswith('~$'):
                print(f'Arquivo {file_to_move} está aberto')
                return False

            shutil.move(current_file_path_origin, current_file_path_target)
            print(f'Arquivo {file_to_move} movido para {current_file_path_target}')
        
        

  
        
class TesteStreamlit:
    def __init__(self, host):
        # Crie uma instância de ConnectPostgresQL usando o host do seu banco de dados PostgreSQL
        self.db_connection = ConnectPostgresQL(host)
        self.session = self.db_connection.Session()


    def check_and_update_orders_streamlit(self, uploaded_file:file_uploader, col, progress_callback=None):
        start = time.time()
        """Método para verificar e atualizar pedidos ausentes no banco de dados"""

        try:
            # Carrega o arquivo fornecido pelo usuário
            extract_df = pd.read_excel(uploaded_file, sheet_name='2-Resultado', engine='openpyxl', header=1)

            # verifica se a coluna "Nome do Cliente" esta presente no indice 1(2ª linha)
            if 'Pedido Faturamento' in extract_df.iloc[1].values:
                extract_df.columns = extract_df.iloc[1]

            # Padroniza o nome da coluna para minúsculas e substitui espaços por underscore
            extract_df.columns = extract_df.columns.str.lower().str.replace(' ', '_').str.replace('.', '') \
                .str.replace('-', '') \
                .str.replace('ç', 'c') \
                .str.replace('cálculo_reajuste', 'calculo_reajuste')  # Remove o acento "´"

            # Verifica se a coluna existe no arquivo
            col_lower = col.lower().replace(' ', '_')
            if col_lower not in extract_df.columns:
                print(f'Coluna {col} não encontrada no arquivo')
                return

            # Converte a coluna PEDIDO para numérico
            extract_df[col_lower] = pd.to_numeric(extract_df[col_lower], errors='coerce')
            print(f'Total de registros no extrator: {len(extract_df)}')

            # Filtra o DataFrame para incluir apenas os pedidos mais recentes
            extract_df = extract_df[extract_df[col_lower] >= 0]

            # Carrega os pedidos já existentes no banco de dados, convertendo a coluna para inteiro
            existing_orders = set(int(order) for order in pd.read_sql_query(
                f'SELECT DISTINCT {col} FROM {OrdersTable.__tablename__}', self.db_connection.engine)[col])

            # Identifica os pedidos ausentes
            new_orders = set(extract_df[col_lower]) - existing_orders
            print(f'Total de novos pedidos no extrator: {len(new_orders)}')

            if len(new_orders) > 0:
                # Reinicializa a variável new_orders_df
                new_orders_df = pd.DataFrame()

                # Cria um DataFrame apenas com os pedidos ausentes
                new_orders_df = extract_df[extract_df[col_lower].isin(new_orders)].copy()

                # Verifica se há novos pedidos antes de continuar
                if not new_orders_df.empty:
                    # caminho do diretório NOVOS_PEDIDOS
                    path = r'/home/administrator/WindowsShare/01 - FATURAMENTO/03 - DATA_RAW'
                    # cria o diretório NOVOS_PEDIDOS se não existir
                    os.makedirs(path, exist_ok=True)
                    # percorre o DataFrame agrupando os pedidos por cliente
                    for order_number, order_group in new_orders_df.groupby(col_lower):
                        # remove caracteres inválidos do nome do cliente e cria o nome do arquivo
                        client_name_valid = order_group['nome_do_cliente'].iloc[0].translate(
                            str.maketrans('', '', r'\/:*?"<>|'))
                        # define o nome e cria o arquivo
                        file_name = f'{order_number}_{client_name_valid}.xlsx'
                        # caminho completo do arquivo para salvar
                        file_path = os.path.join(path, file_name)
                        # salva o arquivo em excel
                        order_group.to_excel(file_path, sheet_name='CONSOLIDADO', index=False, engine='openpyxl')
                        print(f'Novo arquivo {file_name} criado.')

                        # Atualize o progresso
                        if progress_callback is not None:
                            progress_callback((len(new_orders_df) / len(new_orders)) * 100)

                    # Atualiza o banco de dados com os pedidos ausentes
                    print('Atualizando banco de dados....!')
                    try:
                        new_orders_df = new_orders_df.rename(columns={'cálculo_reajuste': 'calculo_reajuste'})

                        new_orders_df.to_sql(OrdersTable.__tablename__, self.db_connection.engine, if_exists='append', index=False, method='multi')
                        print(Fore.GREEN + 'Banco de dados atualizado com novos pedidos' + Fore.RESET)
                    except IntegrityError as e:
                        print('Erro ao atualizar o banco de dados:', e)

                    # pula o processamento dos clientes abaixo (grandes clientes)
                    special_clients = ['teste']

                    def save_order_excel(order, project):
                        order_df = extract_df[extract_df[col_lower] == order]
                        
                        if not order_df.empty:
                            client_name = order_df['CLIENTE'].iloc[0]
                            
                            if client_name in special_clients:
                                print(f'Relatório {client_name} será gerado manualmente')
                                return

                            client_name_safe = re.sub(r'[^a-zA-Z0-9_]', '_', unidecode.unidecode(client_name))
                            sheet_names = ['LAVORO', 'CONSOLIDADO']

                            # itera sobre as sheets do arquivo
                            for sheet in sheet_names:
                                # Define o nome do arquivo no padrão desejado, incluindo o número do projeto
                                file_name = f'{project}_{order}_{client_name_safe}.xlsx'
                                # Caminho completo do arquivo para salvar
                                file_path = os.path.join(path, file_name)
                                # Salva o arquivo em excel
                                order_df.to_excel(file_path, sheet_name=sheet, index=False, engine='openpyxl')
                                print(f'Arquivo {file_name} criado.')

                                # Atualize o progresso
                                if progress_callback is not None:
                                    progress_callback((len(new_orders_df) / len(new_orders)) * 100)
                                    

                    with ThreadPoolExecutor() as executor:
                        executor.map(lambda order: save_order_excel(order, 'projeto'), new_orders)

                    print(f'Pedidos salvos no diretório NOVOS_PEDIDOS')
                    print('Verificação e atualização concluídas.\n')
                    end = time.time()
                    #print(f'Tempo de execução do código: {end - start}')

                else:
                    print('Nenhum pedido novo encontrado.\n')

        except PermissionError as e:
            print(f"Erro de permissão ao acessar o arquivo: {e}")
            print('Corrija as permissões e tente novamente.')
            return False
        except Exception as e:
            print(f"Erro ao processar arquivo: {e}")
            return False

    


    